"""
CSV/Excel Parser & Validator Module for Bulk Imports

Handles parsing and validation of CSV/Excel files for bulk import:
- Student, Instructor, Guardian, Assessment Plan
- Validates required/optional fields
- Generates detailed error reports (which rows, what's wrong)
- Handles large files (1,000+ rows) efficiently

Usage:
    from edupro_sms.import_parser import parse_and_validate_file
    validated_records, errors = parse_and_validate_file('students.xlsx', 'Student')
"""

import frappe
from frappe import _
import csv
import json
from datetime import datetime
from typing import List, Dict, Tuple

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ValidationError:
    """Single validation error for a record"""
    def __init__(self, row_num: int, field: str, value: str, message: str):
        self.row_num = row_num
        self.field = field
        self.value = value
        self.message = message

    def to_dict(self):
        return {
            'row': self.row_num,
            'field': self.field,
            'value': self.value,
            'message': self.message
        }


class RecordValidator:
    """Base validator for all DocTypes"""

    REQUIRED_FIELDS = {}
    OPTIONAL_FIELDS = {}
    FIELD_VALIDATORS = {}

    @classmethod
    def validate_record(cls, record: Dict, row_num: int) -> List[ValidationError]:
        """
        Validate a single record. Returns list of validation errors (empty = valid).

        Args:
            record: Dict with field names as keys
            row_num: Row number in file (for error reporting)

        Returns:
            List of ValidationError objects (empty = valid)
        """
        errors = []

        # Check required fields
        for field in cls.REQUIRED_FIELDS:
            if not record.get(field) or str(record.get(field)).strip() == '':
                errors.append(ValidationError(
                    row_num, field, record.get(field, ''),
                    f"Required field '{field}' is missing or empty"
                ))

        # Run field-specific validators
        for field, validator_func in cls.FIELD_VALIDATORS.items():
            if field in record and record[field]:
                field_errors = validator_func(record[field], row_num, field)
                errors.extend(field_errors)

        return errors

    @staticmethod
    def validate_email(value: str, row_num: int, field: str) -> List[ValidationError]:
        """Validate email format"""
        if not value or '@' not in str(value):
            return [ValidationError(row_num, field, value, "Invalid email format")]
        return []

    @staticmethod
    def validate_date(value: str, row_num: int, field: str) -> List[ValidationError]:
        """Validate date format (YYYY-MM-DD)"""
        try:
            if value:
                datetime.strptime(str(value).strip(), '%Y-%m-%d')
            return []
        except ValueError:
            return [ValidationError(
                row_num, field, value,
                "Invalid date format. Use YYYY-MM-DD (e.g., 2010-05-15)"
            )]

    @staticmethod
    def validate_gender(value: str, row_num: int, field: str) -> List[ValidationError]:
        """Validate gender (M/F)"""
        if value and str(value).upper() not in ['M', 'F']:
            return [ValidationError(row_num, field, value, "Gender must be 'M' or 'F'")]
        return []

    @staticmethod
    def validate_exists(doctype: str, field_name: str):
        """Factory: Create validator that checks if value exists in DocType"""
        def validator(value: str, row_num: int, field: str) -> List[ValidationError]:
            if not value:
                return []
            try:
                count = frappe.db.count(doctype, filters={field_name: str(value).strip()})
                if count == 0:
                    return [ValidationError(
                        row_num, field, value,
                        f"{doctype} with {field_name}='{value}' not found"
                    )]
            except Exception as e:
                frappe.logger().warning(f"Validation check failed: {e}")
            return []
        return validator


class StudentValidator(RecordValidator):
    """Validator for Student import"""

    REQUIRED_FIELDS = [
        'Student Name',
        'Email',
        'Program',
        'Student Group (Class)'
    ]

    OPTIONAL_FIELDS = [
        'Date of Birth',
        'Gender',
        'Boarding Type'
    ]

    FIELD_VALIDATORS = {
        'Email': RecordValidator.validate_email,
        'Date of Birth': RecordValidator.validate_date,
        'Gender': RecordValidator.validate_gender,
    }

    @classmethod
    def validate_record(cls, record: Dict, row_num: int) -> List[ValidationError]:
        errors = super().validate_record(record, row_num)

        # Additional validations
        # Check if email already exists
        if record.get('Email'):
            email = str(record['Email']).strip()
            existing = frappe.db.count('Student', filters={'student_email_id': email})
            if existing > 0:
                errors.append(ValidationError(
                    row_num, 'Email', email,
                    f"Email '{email}' already exists in system"
                ))

        # Check if Program exists
        if record.get('Program'):
            program = str(record['Program']).strip()
            try:
                count = frappe.db.count('Program', filters={'name': program})
                if count == 0:
                    errors.append(ValidationError(
                        row_num, 'Program', program,
                        f"Program '{program}' not found. Create it first."
                    ))
            except Exception:
                pass

        # Check if Student Group (Class) exists
        if record.get('Student Group (Class)'):
            class_name = str(record['Student Group (Class)']).strip()
            try:
                count = frappe.db.count('Student Group', filters={'name': class_name})
                if count == 0:
                    errors.append(ValidationError(
                        row_num, 'Student Group (Class)', class_name,
                        f"Student Group (Class) '{class_name}' not found. Create it first."
                    ))
            except Exception:
                pass

        # Validate boarding type if provided
        if record.get('Boarding Type'):
            boarding = str(record['Boarding Type']).strip()
            if boarding not in ['Day Boarder', 'Full Boarder']:
                errors.append(ValidationError(
                    row_num, 'Boarding Type', boarding,
                    "Boarding Type must be 'Day Boarder' or 'Full Boarder'"
                ))

        return errors


class InstructorValidator(RecordValidator):
    """Validator for Instructor import"""

    REQUIRED_FIELDS = [
        'Instructor Name',
        'Email',
        'Subjects'
    ]

    OPTIONAL_FIELDS = [
        'Class Teacher',
        'Class Teacher For'
    ]

    FIELD_VALIDATORS = {
        'Email': RecordValidator.validate_email,
    }

    @classmethod
    def validate_record(cls, record: Dict, row_num: int) -> List[ValidationError]:
        errors = super().validate_record(record, row_num)

        # Check if email already exists
        if record.get('Email'):
            email = str(record['Email']).strip()
            existing = frappe.db.count('User', filters={'email': email})
            if existing > 0:
                errors.append(ValidationError(
                    row_num, 'Email', email,
                    f"Email '{email}' already exists (User/Instructor)"
                ))

        # Validate Class Teacher field
        if record.get('Class Teacher'):
            class_teacher = str(record['Class Teacher']).strip().lower()
            if class_teacher not in ['yes', 'no']:
                errors.append(ValidationError(
                    row_num, 'Class Teacher', record['Class Teacher'],
                    "Class Teacher must be 'Yes' or 'No'"
                ))

        # If Class Teacher = Yes, then Class Teacher For is required
        if record.get('Class Teacher'):
            if str(record['Class Teacher']).strip().lower() == 'yes':
                if not record.get('Class Teacher For'):
                    errors.append(ValidationError(
                        row_num, 'Class Teacher For', '',
                        "Class Teacher For is required when Class Teacher = Yes"
                    ))

        return errors


class GuardianValidator(RecordValidator):
    """Validator for Guardian import"""

    REQUIRED_FIELDS = [
        'Guardian Name',
        'Email',
        'Student IDs'
    ]

    OPTIONAL_FIELDS = []

    FIELD_VALIDATORS = {
        'Email': RecordValidator.validate_email,
    }

    @classmethod
    def validate_record(cls, record: Dict, row_num: int) -> List[ValidationError]:
        errors = super().validate_record(record, row_num)

        # Check if email already exists
        if record.get('Email'):
            email = str(record['Email']).strip()
            existing = frappe.db.count('Guardian', filters={'email_address': email})
            if existing > 0:
                errors.append(ValidationError(
                    row_num, 'Email', email,
                    f"Guardian email '{email}' already exists"
                ))

        # Parse and validate Student IDs (comma-separated)
        if record.get('Student IDs'):
            student_ids = [s.strip() for s in str(record['Student IDs']).split(',')]
            for student_id in student_ids:
                if student_id:
                    try:
                        count = frappe.db.count('Student', filters={'name': student_id})
                        if count == 0:
                            errors.append(ValidationError(
                                row_num, 'Student IDs', student_id,
                                f"Student ID '{student_id}' not found"
                            ))
                    except Exception:
                        pass

        return errors


class AssessmentPlanValidator(RecordValidator):
    """Validator for Assessment Plan import"""

    REQUIRED_FIELDS = [
        'Class (Student Group)',
        'Subject (Course)',
        'Academic Term',
        'Exam Name',
        'Schedule Date'
    ]

    OPTIONAL_FIELDS = [
        'Start Time',
        'End Time',
        'Criteria'
    ]

    FIELD_VALIDATORS = {
        'Schedule Date': RecordValidator.validate_date,
    }

    @classmethod
    def validate_record(cls, record: Dict, row_num: int) -> List[ValidationError]:
        errors = super().validate_record(record, row_num)

        # Check if Student Group exists
        if record.get('Class (Student Group)'):
            class_name = str(record['Class (Student Group)']).strip()
            try:
                count = frappe.db.count('Student Group', filters={'name': class_name})
                if count == 0:
                    errors.append(ValidationError(
                        row_num, 'Class (Student Group)', class_name,
                        f"Student Group '{class_name}' not found"
                    ))
            except Exception:
                pass

        # Check if Course exists
        if record.get('Subject (Course)'):
            subject = str(record['Subject (Course)']).strip()
            try:
                count = frappe.db.count('Course', filters={'name': subject})
                if count == 0:
                    errors.append(ValidationError(
                        row_num, 'Subject (Course)', subject,
                        f"Course '{subject}' not found"
                    ))
            except Exception:
                pass

        # Check if Academic Term exists
        if record.get('Academic Term'):
            term = str(record['Academic Term']).strip()
            try:
                count = frappe.db.count('Academic Term', filters={'name': term})
                if count == 0:
                    errors.append(ValidationError(
                        row_num, 'Academic Term', term,
                        f"Academic Term '{term}' not found"
                    ))
            except Exception:
                pass

        return errors


def parse_csv_file(file_path: str) -> Dict:
    """
    Parse CSV file. Returns dict with:
    - headers: List of column names
    - rows: List of dicts (one per row)
    - errors: List of parsing errors (if any)
    """
    records = []
    headers = []
    parse_errors = []

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []

            for row_num, row in enumerate(reader, start=2):  # Start at 2 (after header)
                records.append((row_num, dict(row)))

    except Exception as e:
        parse_errors.append(f"CSV parsing error: {str(e)}")

    return {
        'headers': headers,
        'records': records,
        'errors': parse_errors
    }


def parse_excel_file(file_path: str, sheet_name: str = None) -> Dict:
    """
    Parse Excel file. Returns dict with:
    - headers: List of column names
    - rows: List of dicts (one per row)
    - errors: List of parsing errors (if any)
    """
    if not HAS_OPENPYXL:
        return {
            'headers': [],
            'records': [],
            'errors': ["openpyxl not installed. Run: pip install openpyxl"]
        }

    records = []
    headers = []
    parse_errors = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Use first sheet or specified sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Skip "Instructions" sheet if it exists
            sheet_to_use = [s for s in wb.sheetnames if s != 'Instructions'][0] \
                if any(s != 'Instructions' for s in wb.sheetnames) \
                else wb.active
            ws = wb[sheet_to_use]

        # Extract headers from first row
        header_row = []
        for cell in ws[1]:
            header_row.append(cell.value)
        headers = header_row

        # Extract data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Convert cells to values
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_data[headers[col_idx]] = cell.value

            # Skip completely empty rows
            if any(v for v in row_data.values() if v):
                records.append((row_num, row_data))

    except Exception as e:
        parse_errors.append(f"Excel parsing error: {str(e)}")

    return {
        'headers': headers,
        'records': records,
        'errors': parse_errors
    }


def _resolve_uploaded_file_path(file_url: str) -> str:
    """file_path here is actually a Frappe file_url (e.g. /private/files/
    students.csv) returned by /api/method/upload_file -- resolve it to a
    real filesystem path that csv/openpyxl can open() directly."""
    file_doc = frappe.get_doc('File', {'file_url': file_url})
    return file_doc.get_full_path()


@frappe.whitelist()
def parse_and_validate_file(file_path: str, doctype: str, sheet_name: str = None) -> Dict:
    """
    Parse and validate a CSV/Excel file. Returns validation results.

    Args:
        file_path: file_url of an already-uploaded file (from
            /api/method/upload_file), e.g. "/private/files/students.csv"
        doctype: 'Student', 'Instructor', 'Guardian', or 'Assessment Plan'
        sheet_name: Optional, for Excel files (defaults to first non-Instructions sheet)

    Returns:
        Dict with:
        - valid_records: List of dicts (ready to import)
        - invalid_records: List of dicts with errors
        - error_summary: Summary stats
        - errors: List of ValidationError dicts
    """

    real_path = _resolve_uploaded_file_path(file_path)

    # Determine parser & validator
    if file_path.endswith('.csv'):
        parse_result = parse_csv_file(real_path)
    elif file_path.endswith(('.xlsx', '.xls')):
        parse_result = parse_excel_file(real_path, sheet_name)
    else:
        return {
            'valid_records': [],
            'invalid_records': [],
            'error_summary': {'parse_errors': 1, 'validation_errors': 0},
            'errors': [{'message': 'Unsupported file format. Use .csv or .xlsx'}]
        }

    # Check parsing errors
    if parse_result['errors']:
        return {
            'valid_records': [],
            'invalid_records': [],
            'error_summary': {'parse_errors': len(parse_result['errors']), 'validation_errors': 0},
            'errors': [{'message': e} for e in parse_result['errors']]
        }

    # Map doctype to validator
    validator_map = {
        'Student': StudentValidator,
        'Instructor': InstructorValidator,
        'Guardian': GuardianValidator,
        'Assessment Plan': AssessmentPlanValidator,
    }

    validator = validator_map.get(doctype)
    if not validator:
        return {
            'valid_records': [],
            'invalid_records': [],
            'error_summary': {'parse_errors': 1, 'validation_errors': 0},
            'errors': [{'message': f'Unknown doctype: {doctype}'}]
        }

    # Validate each record
    valid_records = []
    invalid_records = []
    all_errors = []

    for row_num, record in parse_result['records']:
        errors = validator.validate_record(record, row_num)

        if errors:
            invalid_records.append({
                'row_num': row_num,
                'data': record,
                'errors': [e.to_dict() for e in errors]
            })
            all_errors.extend(errors)
        else:
            valid_records.append({
                'row_num': row_num,
                'data': record
            })

    return {
        'total_records': len(parse_result['records']),
        'valid_records': valid_records,
        'invalid_records': invalid_records,
        'error_summary': {
            'parse_errors': 0,
            'validation_errors': len(all_errors),
            'valid_count': len(valid_records),
            'invalid_count': len(invalid_records)
        },
        'errors': [e.to_dict() for e in all_errors]
    }
