"""
Template Generator Module for Bulk Imports

Generates pre-formatted Excel/CSV templates for different DocTypes:
- Student: name, email, DOB, gender, ID, boarding_type, program, student_group
- Instructor: name, email, subjects (comma-separated), class_teacher_flag
- Guardian: name, email, student_ids (comma-separated)
- Assessment Plan: student_group, course, academic_term, schedule_date, criteria

Usage:
    from edupro_sms.template_generator import generate_student_template
    file_path = generate_student_template()  # Returns path to Excel file
"""

import frappe
from frappe import _
from datetime import datetime
import os
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_student_template():
    """
    Generate Excel template for Student bulk import

    Headers: name, email, dob, gender, admission_number, boarding_type, program, student_group

    Returns:
        str: Path to generated Excel file
    """
    if not HAS_OPENPYXL:
        frappe.throw("openpyxl library not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Define headers
    headers = [
        "Student Name",          # A
        "Email",                 # B
        "Date of Birth",         # C (YYYY-MM-DD format)
        "Gender",                # D (M/F)
        "Boarding Type",         # E (Day Boarder/Full Boarder)
        "Program",               # F (e.g., "IGCSE Science")
        "Student Group (Class)"  # G (e.g., "Form 1A")
    ]

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Add sample data rows (3 examples)
    sample_data = [
        ["John Doe", "john.doe@school.edu", "2010-05-15", "M", "Day Boarder", "IGCSE Science", "Form 1A"],
        ["Jane Smith", "jane.smith@school.edu", "2010-08-22", "F", "Full Boarder", "IGCSE Science", "Form 1A"],
        ["Bob Wilson", "bob.wilson@school.edu", "2010-03-10", "M", "Day Boarder", "IGCSE Commerce", "Form 1B"],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set column widths
    column_widths = [20, 25, 18, 10, 18, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        "BULK STUDENT IMPORT INSTRUCTIONS",
        "",
        "1. Fill in the 'Students' sheet with your student data",
        "2. Required fields: Student Name, Email, Program, Student Group",
        "3. Optional fields: Date of Birth, Gender, Boarding Type (defaults to 'Day Boarder')",
        "4. Date format: YYYY-MM-DD (e.g., 2010-05-15)",
        "5. Gender: M or F",
        "6. Boarding Type: 'Day Boarder' or 'Full Boarder'",
        "7. Program must exist (e.g., 'IGCSE Science')",
        "8. Student Group must exist (e.g., 'Form 1A')",
        "9. Email must be unique (not already in system)",
        "10. Do NOT edit the header row",
        "",
        "Sample data is provided as a guide. Delete or replace it with your own data.",
        "",
        "Upload completed file using the 'Import Data' page in Edupro SMS.",
    ]

    for row_num, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row_num, column=1)
        cell.value = instruction
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    instructions_ws.column_dimensions['A'].width = 80

    # Save to temp file
    file_path = f"/tmp/student_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)

    return file_path


def generate_instructor_template():
    """
    Generate Excel template for Instructor bulk import

    Headers: name, email, subjects (comma-separated), class_teacher_flag

    Returns:
        str: Path to generated Excel file
    """
    if not HAS_OPENPYXL:
        frappe.throw("openpyxl library not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instructors"

    headers = [
        "Instructor Name",       # A
        "Email",                 # B
        "Subjects",              # C (comma-separated, e.g., "Mathematics, Physics")
        "Class Teacher",         # D (Yes/No)
        "Class Teacher For"      # E (if Yes above, which class? e.g., "Form 1A")
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_data = [
        ["Grace Mensah", "grace.mensah@school.edu", "Mathematics, Physics", "Yes", "Form 1A"],
        ["Kwame Boateng", "kwame.boateng@school.edu", "English, History", "Yes", "Form 1B"],
        ["Ama Asante", "ama.asante@school.edu", "Biology, Chemistry", "No", ""],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    column_widths = [20, 25, 30, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    file_path = f"/tmp/instructor_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)

    return file_path


def generate_guardian_template():
    """
    Generate Excel template for Guardian bulk import

    Headers: name, email, student_ids (comma-separated)

    Returns:
        str: Path to generated Excel file
    """
    if not HAS_OPENPYXL:
        frappe.throw("openpyxl library not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guardians"

    headers = [
        "Guardian Name",    # A
        "Email",           # B
        "Student IDs",     # C (comma-separated, e.g., "STU001, STU002")
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_data = [
        ["Mary Owusu", "mary.owusu@email.com", "STU001, STU002"],
        ["Robert Mensah", "robert.mensah@email.com", "STU003"],
        ["Abena Asempa", "abena.asempa@email.com", "STU004, STU005, STU006"],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    column_widths = [20, 25, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    file_path = f"/tmp/guardian_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)

    return file_path


def generate_assessment_plan_template():
    """
    Generate Excel template for Assessment Plan bulk import

    Headers: student_group, course, academic_term, schedule_date, criteria

    Returns:
        str: Path to generated Excel file
    """
    if not HAS_OPENPYXL:
        frappe.throw("openpyxl library not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Plans"

    headers = [
        "Class (Student Group)",  # A
        "Subject (Course)",       # B
        "Academic Term",          # C
        "Exam Name",              # D
        "Schedule Date",          # E (YYYY-MM-DD)
        "Start Time",             # F (HH:MM)
        "End Time",               # G (HH:MM)
        "Criteria",               # H (e.g., "Term Mark, Exam Mark")
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_data = [
        ["Form 1A", "Mathematics", "2026 (Term 1)", "Term 1 Assessment", "2026-04-15", "09:00", "11:00", "Term Mark, Exam Mark"],
        ["Form 1A", "Physics", "2026 (Term 1)", "Term 1 Assessment", "2026-04-16", "09:00", "11:00", "Term Mark, Exam Mark"],
        ["Form 1B", "Mathematics", "2026 (Term 1)", "Term 1 Assessment", "2026-04-17", "09:00", "11:00", "Term Mark, Exam Mark"],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    column_widths = [18, 18, 18, 20, 18, 12, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    file_path = f"/tmp/assessment_plan_import_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)

    return file_path


@frappe.whitelist()
def download_template(template_type):
    """
    Whitelisted function to download a template. Hit directly via GET
    (window.location.href = '/api/method/...') -- setting
    frappe.response.filecontent/filename/type is what actually triggers
    a browser download; returning a plain dict (the previous
    implementation) just renders as JSON in the browser.

    Args:
        template_type: 'student', 'instructor', 'guardian', or 'assessment_plan'
    """
    if template_type == "student":
        file_path = generate_student_template()
    elif template_type == "instructor":
        file_path = generate_instructor_template()
    elif template_type == "guardian":
        file_path = generate_guardian_template()
    elif template_type == "assessment_plan":
        file_path = generate_assessment_plan_template()
    else:
        frappe.throw(f"Invalid template type: {template_type}")

    with open(file_path, 'rb') as f:
        content = f.read()
    os.remove(file_path)

    frappe.local.response.filename = os.path.basename(file_path)
    frappe.local.response.filecontent = content
    frappe.local.response.type = "download"
